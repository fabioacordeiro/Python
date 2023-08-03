# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

# pip install openpyxl

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
# print(ROOT_FOLDER)
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'
print(WORKBOOK_PATH)
workbook = Workbook()
worksheet: Worksheet = workbook.active

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')
workbook.save(WORKBOOK_PATH)

lista = [
    # nome,....idade, nota
    ['Joao', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10]
]
'''
for l in range(1, 5):
    for c in range(1, 4):
        print(f'Linha:{l} > Coluna:{c}')
'''
for lista_row in lista:
    print(lista_row)
coluna = 0
for n, lista_row in enumerate(lista, start=2):
    print(f'Linha:{n}')
    for valor in lista_row:
        coluna += 1
        print(f'Valor de Linha:{valor} coluna:{coluna}')
    coluna = 0

print('Arquivo criado e Salvo')
