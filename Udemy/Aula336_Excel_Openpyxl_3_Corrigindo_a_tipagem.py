# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell  # Aqui eu corrijo a tipagem de Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook1.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome para a planilha
sheet_name = 'Minha planilha'

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]  # inserindo a tipagem
# onde 2 é o número da linha que eu quero iniciar pode informar a linha final a coluna inicial e final com max_row, min_col, max_col, etc
for row in worksheet.iter_rows(min_row=2, ):
    for cell in row:
        # agora quando colocamos cell. podemos visualizar todas as opções de cell
        print(cell.value, end='\t')
    print()

print()
print(worksheet['b3'])
print(worksheet['b3'].value)

workbook.save(WORKBOOK_PATH)
