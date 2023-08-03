# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# openpyxl - manipulando as planilhas do Workbook
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas

# pip install openpyxl

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
# print(ROOT_FOLDER)
WORKBOOK_PATH = ROOT_FOLDER / 'workbook2.xlsx'
print(WORKBOOK_PATH)
workbook = Workbook()
worksheet: Worksheet = workbook.active

# Nome para a planilha
sheet_name = 'Minha planilha'
# Criamos a planilha
workbook.create_sheet(sheet_name, 0)
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])


# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')


nomes = [
    # nome,....idade, nota
    ['Joao', 14.0, 5.5],
    ['Maria', 13.0, 9.7],
    ['Luiz', 15.0, 8.8],
    ['Alberto', 16.0, 10.0]
]

for nome in nomes:
    worksheet.append(nome)

workbook.save(WORKBOOK_PATH)
